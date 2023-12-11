from types import NoneType
from fastapi import APIRouter, File, UploadFile, Header
import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from io import BytesIO
from unicodedata import normalize
import datetime
import re

from api.schemas.person import Teacher
from api.schemas.timeslot import TimeslotBase, Timeslot
from api.cruds.teacher import TeacherRepo
from api.cruds.timeslot import TimeslotRepo, YearMonth
from api.cruds.meta import MetaRepo

from api.myutils.utilfunc import excel_date_to_datetime, str2int_timeslot_num, get_start_end_time
from api.myutils.const import CellBlock, class_times


router = APIRouter()

# xlsx file upload
@router.post("/calc_salary/upload_class_sheet")
async def upload_class_sheet(year: int, month: int, file: UploadFile = File(...), school_id: str = Header(...)):
    if file.filename == None:
        raise Exception("File name is None")
    
    if not file.filename.endswith(".xlsx"): #type: ignore
        raise Exception("File type not supported")

    wb = xl.load_workbook(filename=BytesIO(await file.read()), data_only=True)
    make_timeslots(wb, school_id, year, month)
    

@router.get("/calc_salary/calculate")
async def calculate_salary(year: int, month: int | None = None, school_id: str = Header(...)):
    teacher_list = TeacherRepo.list(school_id, YearMonth(year=year, month=month))
    timeslot_list = TimeslotRepo.list(school_id, YearMonth(year=year, month=month))
    

def make_timeslots(
        wb: xl.Workbook, 
        school_id: str,
        year: int,
        month: int | None = None,
    ) -> list[Timeslot]:
    
    teacher_list = TeacherRepo.list(school_id)
    if month != None:
        [TeacherRepo.update(teacher.id, teacher.get_base(), YearMonth(year=year, month=month)) for teacher in teacher_list]
        ws_list = [wb[ws_name] for ws_name in wb.sheetnames if normalize("NFKC" , ws_name).startswith(f"{month}月")]
    else:
        ws_list = [wb[ws_name] for ws_name in wb.sheetnames]
    
    teacher_dict = {teacher.display_name: teacher.id for teacher in teacher_list}
    
    timeslot_list = []
    for ws in ws_list:
        date = None
        for i in range(CellBlock.MAX_BLOCKS_ROW):
            date_cell: datetime.date | int | None = ws.cell(
                row = i * CellBlock.BLOCK_SIZE + CellBlock.INFO_DATE_IDX, 
                column = CellBlock.INFO_COL
            ).value # type: ignore

            timeslot_num_cell: str | int | None = ws.cell(
                row = i * CellBlock.BLOCK_SIZE + CellBlock.INFO_TIMESLOTNUM_IDX, 
                column = CellBlock.INFO_COL
            ).value # type: ignore

            time_cell: str | None = ws.cell(
                row = i * CellBlock.BLOCK_SIZE + CellBlock.INFO_TIME_IDX, 
                column = CellBlock.INFO_COL
            ).value # type: ignore

            if type(date_cell) not in [datetime.date, int, NoneType]:
                continue
            
            if type(timeslot_num_cell) not in [str, int, NoneType]:
                raise Exception(f"timeslot_num_cell type is {type(timeslot_num_cell)}")
            
            if type(time_cell) not in [str, NoneType]:
                raise Exception(f"time_cell type is {type(time_cell)}")
            
            
            if (timeslot_num_cell == None) | (time_cell == None):
                continue

            # 日付を更新
            if date_cell != None:
                date = normalize_date(date_cell) # type: ignore
            
            # 更新後の日付がNoneの場合はスキップ
            if date == None:
                continue

            if date.year != year: # type: ignore
                continue
             
            if (month != None) & (date.month != month): # type: ignore
                continue

            timeslot_num = get_timeslot_num(time_cell) # type: ignore
            start_time, end_time = normalize_time(date, time_cell) # type: ignore

            # timeslot_baseを作成
            timeslot_base_lecture = TimeslotBase(
                school_id=school_id,
                date=date, # type: ignore
                timeslot_number=timeslot_num,
                timeslot_type="lecture",
                start_time=start_time,
                end_time=end_time
            )

            for j in range(CellBlock.MAX_BLOCKS_COL):
                display_name: str | None = ws.cell(
                    row = i * CellBlock.BLOCK_SIZE + CellBlock.BLOCK_NAME_IDX, 
                    column = CellBlock.INFO_COL + 1 + j
                ).value # type: ignore

                cell1: str | None = ws.cell(
                    row = i * CellBlock.BLOCK_SIZE + CellBlock.BLOCK_CELL1_IDX, 
                    column = CellBlock.INFO_COL + 1 + j
                ).value # type: ignore

                cell2: str | None = ws.cell(
                    row = i * CellBlock.BLOCK_SIZE + CellBlock.BLOCK_CELL2_IDX, 
                    column = CellBlock.INFO_COL + 1 + j
                ).value # type: ignore

                if type(display_name) not in [str, NoneType]:
                    raise Exception(f"display_name type is {type(display_name)}")
                
                if type(cell1) not in [str, NoneType]:
                    raise Exception(f"cell1 type is {type(cell1)}")
                
                if type(cell2) not in [str, NoneType]:
                    raise Exception(f"cell2 type is {type(cell2)}")
                           
                # 講師名がNoneであれば無視
                if display_name == None:
                    continue

                # cellが二つともNoneであれば無視
                if (cell1 == None) and (cell2 == None):
                    continue

                # 講師名がteacher_dictになければ無視
                if display_name not in teacher_dict:
                    continue

                teacher_id = teacher_dict[display_name]
                officework_end_time = get_officework_end_time(start_time, end_time, cell1, cell2)
                if officework_end_time != None:
                    timeslot_base_office = TimeslotBase(
                        school_id=school_id,
                        date=date, # type: ignore
                        timeslot_number=0,
                        timeslot_type="office_work",
                        start_time=start_time,
                        end_time=officework_end_time # type: ignore
                    )
                    timeslot = TimeslotRepo.create(teacher_id, timeslot_base_office)
                else:
                    timeslot = TimeslotRepo.create(teacher_id, timeslot_base_lecture)          
                
                timeslot_list.append(timeslot)
    return timeslot_list

                
def normalize_timeslot_num(timeslot_num_cell: str | int) -> int:
    if type(timeslot_num_cell) == str:
        return str2int_timeslot_num(timeslot_num_cell)
    elif type(timeslot_num_cell) == int:
        return timeslot_num_cell
    else:
        raise Exception(f"timeslot_num_cell type is {type(timeslot_num_cell)}")

def get_timeslot_num(time_cell: str) -> int:
    time_cell = normalize("NFKC", time_cell)
    return class_times[time_cell]
    
def normalize_time(date: datetime.date, time_cell: str) -> tuple[datetime.datetime, datetime.datetime]:
    time_cell = normalize("NFKC", time_cell)
    start_time_str, end_time_str = time_cell.split("-")
    start_time_time = datetime.datetime.strptime(start_time_str, "%H:%M").time()
    end_time_time = datetime.datetime.strptime(end_time_str, "%H:%M").time()
    start_time = datetime.datetime.combine(date, start_time_time)
    end_time = datetime.datetime.combine(date, end_time_time)
    return start_time, end_time

def normalize_date(date_cell: datetime.date | int) -> datetime.date:
    if type(date_cell) == datetime.date:
        return date_cell
    elif type(date_cell) == int:
        return excel_date_to_datetime(date_cell)
    else:
        raise Exception(f"date_cell type is {type(date_cell)}")

# cell1, cell2のどちらかが"事務"を含む場合、その時間を分単位で返す
def get_officework_end_time(start_time: datetime.datetime, end_time: datetime.datetime, cell1: str | None, cell2: str | None) -> datetime.datetime | None:
    def is_officework_cell(cell: str | None) -> datetime.datetime | None:
        if cell == None:
            return None
        cell = normalize("NFKC", cell) # type: ignore
        # cellが"事務"を含む場合
        if "事務" in cell:
            result = re.findall(r"事務(\d+)", cell)
            if len(result) == 0:
                return end_time
            else:
                td = int(result[0])
                return start_time + datetime.timedelta(minutes=td)
        else:
            return None
        
    cell1_officework_time = is_officework_cell(cell1)
    cell2_officework_time = is_officework_cell(cell2)

    if cell1_officework_time != None:
        return cell1_officework_time
    else:
        return cell2_officework_time